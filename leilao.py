from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import load_workbook
import json

wb = load_workbook(r"D:\Municipios.xlsx")
ws = wb.active
linhaf = ws.max_row
lista_link = []
leiloes = []
driver = webdriver.Edge()

driver.get('https://www.grupoarremateleiloes.com.br/busca/#Engine=Start&Pagina=1&Busca=&Mapa=&ID_Categoria=1&PaginaIndex=1&QtdPorPagina=24')

driver.maximize_window()

driver.find_element(By.XPATH,'/html/body/div[2]/div/div/a').click()
pag = 2
while True:
    try:
        while True:
            try:
                driver.find_element(By.XPATH,'//*[@id="ResultadoPaginacao"]/div/nav')
                break
            except:
                sleep(2)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                driver.execute_script("window.scrollBy(0, -100);")


        links = WebDriverWait(driver,60).until(
        EC.presence_of_all_elements_located((By.XPATH,'//*/div[@class="row jsLeiloes"]/div/div/div[@class="dg-leiloes-lista-img"]/a'))
        )
        

        for link in links:
            lista_link.append(link.get_attribute('href'))
        #print(lista_link)
        #print(len(lista_link))
        driver.find_element(By.XPATH,'//*[@id="ResultadoPaginacao"]/div/nav/a[@onclick="BuscaPaginacao({})"]'.format(pag)).click()
        pag=pag+1
    except:
        break
        
driver.execute_script("window.open('');") 
driver.switch_to.window(driver.window_handles[-1])

for link in lista_link: 
    
    driver.get(link)
    titulo = driver.find_element(By.XPATH,'/html/body/main/div/div[1]/div[1]/h1').text.split("|")[0]
    try:
        endereco = driver.find_element(By.XPATH,'/html/body/section[4]/div/div[2]/div/div/div').text
        state = endereco.split(" - ")[-1]
    except:
        state = "not found"
        pass
    
    city = endereco.split(" - ")[-2]
    city = city+" ({})".format(state)
    c=0
    
    for  c in range(linhaf):
        mun = ws.cell(row = c+1, column = 1).value
        
        if mun in city:
            city = mun
            #print(city+"    "+mun)
            
            continue
    valor = driver.find_element(By.XPATH,'//*/strong[@class="ValorAvaliacao"]').text
    
    codigo_leilao = driver.find_element(By.XPATH,'//*/span[@class="dg-lote-titulo-codigo"]/strong').text
    
    descricao = driver.find_element(By.XPATH,'//*/div[@class="dg-lote-descricao-txt"]').text.replace('\n',' ')
    
    try:
        
        imagem = driver.find_element(By.XPATH,'//*/div[2]/div[1]/div[1]/div[1]/div/div/a[2]/img').get_attribute('src')
        
    except:
        
        imagem = driver.find_element(By.XPATH,'//*/div[2]/div[1]/div[1]/div[1]/div/div/a/img').get_attribute('src')
        
    produto = {
            'title': titulo,
            'state': state,
            'city': city,
            'process':codigo_leilao,
            'link':link,
            'address':endereco,
            'description':descricao,
            'image':imagem
        }
    leiloes.append(produto)
    #print(descricao)
    
with open('teste.json', "w") as arquivo:
    json.dump(leiloes, arquivo, indent=4,ensure_ascii=False)       
            
    